from pathlib import Path
import openpyxl

from sqlalchemy.orm import Session

try:
    from .database import engine, SessionLocal
    from . import models
except ImportError:
    from database import engine, SessionLocal
    import models

models.Base.metadata.drop_all(bind=engine)
models.Base.metadata.create_all(bind=engine)

# Each entry: (xlsx filename, Subject name, Topic name)
# Place the xlsx files in a "quiz_data" folder next to this seed.py before running it.
DATA_DIR = Path(__file__).resolve().parent / "quiz_data"
QUIZ_BANKS = [
    ("Physics_Quiz.xlsx", "Physics", "Mechanics"),
    ("Kidney_Quiz.xlsx", "Biology", "Excretory System"),
    ("Titration_Quiz.xlsx", "Chemistry", "Acid-Base Titration"),
]


def _load_bank(db, filename: str, subject_name: str, topic_name: str):
    """Reads one quiz-bank xlsx (Topic, Video Link, Difficulty, Question, OptionA-D, Correct Answer)
    and creates Subject -> Topic -> Subtopic (= xlsx 'Topic' column, i.e. one mini-tutorial) -> Questions."""
    path = DATA_DIR / filename
    if not path.exists():
        print(f"  [skip] {filename} not found in {DATA_DIR}")
        return

    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))[1:]  # drop header row

    subject = db.query(models.Subject).filter(models.Subject.name == subject_name).first()
    if not subject:
        subject = models.Subject(name=subject_name)
        db.add(subject)
        db.commit()

    topic = db.query(models.Topic).filter(
        models.Topic.name == topic_name, models.Topic.subject_id == subject.id
    ).first()
    if not topic:
        topic = models.Topic(name=topic_name, subject_id=subject.id)
        db.add(topic)
        db.commit()

    subtopic_cache = {}  # mini-tutorial name -> Subtopic row
    added = 0
    for row in rows:
        if not row or row[0] is None:
            continue
        mini_tut_name, video_link, difficulty, question, a, b, c, d, correct = row[:9]

        subtopic = subtopic_cache.get(mini_tut_name)
        if subtopic is None:
            subtopic = db.query(models.Subtopic).filter(
                models.Subtopic.name == mini_tut_name, models.Subtopic.topic_id == topic.id
            ).first()
            if not subtopic:
                subtopic = models.Subtopic(name=mini_tut_name, topic_id=topic.id, video_link=video_link)
                db.add(subtopic)
                db.commit()
            subtopic_cache[mini_tut_name] = subtopic

        db.add(models.Question(
            subtopic_id=subtopic.id,
            difficulty=(difficulty or "medium").strip().lower(),
            question=question,
            optionA=a, optionB=b, optionC=c, optionD=d,
            correct_answer=(correct or "A").strip().upper(),
        ))
        added += 1

    db.commit()
    print(f"  [ok] {filename}: {len(subtopic_cache)} mini-tuts, {added} questions")


def seed_db():
    db = SessionLocal()
    print("Seeding database from quiz banks...")
    for filename, subject_name, topic_name in QUIZ_BANKS:
        _load_bank(db, filename, subject_name, topic_name)
    db.close()
    print("Database seeding complete!")


if __name__ == "__main__":
    seed_db()
